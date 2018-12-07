# _*_ coding:utf-8 _*_
import openpyxl
import xlrd
import os
from spider_huawa.huawa_spider import load_page
from spider_huawa.huawa_spider import get_ip_list
import re
from common.img_to_string import img_to_string
import time


# 写入到一个excel
def write_excel(list):
    out_wb = openpyxl.Workbook()  # 打开一个将写的文件
    out_ws = out_wb.create_sheet(index=0)  # 在将写的文件创建sheet
    # ok_list = []
    print(len(list))
    time.sleep(10)
    i = 1
    for row in range(len(list)):
        print("%s 正在导入：%s" % (i, list[row]))
        i = i+1
        # if len([x for x in ok_list if x == list[row][0].value]):
        #     break

        for col in range(len(list[row])):
            if col == 5:
                list[row][5].value = list[row][5].value.replace("\x08", "")
            if col == 6:
                if list[row][6].value == "":
                    try:
                        list[row][6].value = get_phone_by_id(list[row][0].value)
                    except BaseException as err:
                        print(err)
                    finally:
                        print(list[row][6].value)

            out_ws.cell(row + 1, col + 1).value = list[row][col].value  # 写文件
        # ok_list.append(list[row][0].value)
    total_execl_path = r"C:\Users\pig\Desktop\total_1.2.xlsx"
    out_wb.save(total_execl_path)  # 一定要记得保存


# 从花店详情读取电话号码
def get_phone_by_id(store_id):
    url = 'http://www.huawa.com/shop/' + str(store_id)
    html = load_page(url)
    pattern = re.compile(r'.*?花店电话.*?www.huawa.com/phone/(.*?)"', re.S)
    items = re.findall(pattern, html)
    img_url = ""
    if len(items):
        img_url = "http://www.huawa.com/phone/" + items[0]
    return img_to_string(img_url)


if __name__ == "__main__":
    get_ip_list()

    data_path = r'C:\Users\pig\Desktop\data_huawa_v1.1'
    all_data = []

    for dir_path, dir_names, file_names in os.walk(data_path):
        for file_path in file_names:
            # excel文件路径
            excel_path = os.path.join(dir_path, file_path)
            print(excel_path)
            # 打开excel
            excel_now = xlrd.open_workbook(excel_path)
            # 获取第一个sheet内容
            sheet_content = excel_now.sheet_by_index(0)
            # 获取最大行数和列数
            rows = sheet_content.nrows
            cols = sheet_content.ncols
            for i in range(1, rows):
                rows_temp = []
                for j in range(0, cols):
                    rows_temp.append(sheet_content.cell(i, j))
                print(rows_temp)
                all_data.append(rows_temp)
    write_excel(all_data)
